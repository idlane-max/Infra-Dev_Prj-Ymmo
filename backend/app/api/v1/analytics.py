# backend/app/api/v1/analytics.py
"""
Routes d'analytics pour le Dashboard Ymmo.
Utilise Pandas pour les analyses statistiques et prévisions.
Accès réservé aux rôles Direction et Commercial.
"""
from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import List, Dict, Any
from datetime import datetime, timedelta
import pandas as pd
import numpy as np

from app.core.database import get_db
from app.models import core_models
from app.core.security import get_current_user

router = APIRouter()


def _require_staff(current_user: core_models.User):
    """Vérifie que l'utilisateur est Direction ou Commercial"""
    if current_user.role.name not in ["Direction", "Commercial"]:
        raise HTTPException(status_code=403, detail="Accès réservé au personnel Ymmo.")


# ---------------------------------------------------------------------------
# SUMMARY — KPIs globaux
# ---------------------------------------------------------------------------
@router.get("/summary")
def get_summary(
    db: Session = Depends(get_db),
    current_user: core_models.User = Depends(get_current_user),
) -> Dict[str, Any]:
    """
    Retourne les indicateurs clés (KPIs) pour le dashboard.
    Direction : vue globale. Commercial : vue agence uniquement.
    """
    _require_staff(current_user)

    # Récupération des données brutes
    prop_query = db.query(core_models.Property)
    txn_query = db.query(core_models.Transaction)

    if current_user.role.name == "Commercial":
        # Filtrer sur l'agence du commercial
        prop_query = prop_query.join(core_models.User).filter(
            core_models.User.agency_id == current_user.agency_id
        )
        txn_query = txn_query.join(
            core_models.Property, core_models.Transaction.property_id == core_models.Property.id
        ).join(core_models.User, core_models.Property.agent_id == core_models.User.id).filter(
            core_models.User.agency_id == current_user.agency_id
        )

    properties = prop_query.all()
    transactions = txn_query.all()

    # Analyse Pandas
    if properties:
        df_props = pd.DataFrame([{
            "id": p.id,
            "status": p.status.value,
            "price": p.price,
            "type": p.type.value,
            "city": p.city,
            "area": p.area,
            "transaction_mode": p.transaction_mode.value,
            "created_at": p.created_at,
        } for p in properties])
    else:
        df_props = pd.DataFrame(columns=["id", "status", "price", "type", "city", "area", "transaction_mode", "created_at"])

    if transactions:
        df_txns = pd.DataFrame([{
            "id": t.id,
            "price_sold": t.price_sold,
            "transaction_date": t.transaction_date,
        } for t in transactions])
    else:
        df_txns = pd.DataFrame(columns=["id", "price_sold", "transaction_date"])

    total_biens = len(df_props)
    biens_disponibles = len(df_props[df_props["status"] == "available"]) if not df_props.empty else 0
    biens_vendus = len(df_props[df_props["status"] == "sold"]) if not df_props.empty else 0
    biens_en_cours = len(df_props[df_props["status"] == "under_offer"]) if not df_props.empty else 0

    # CA total des transactions
    ca_total = float(df_txns["price_sold"].sum()) if not df_txns.empty else 0.0
    prix_moyen = float(df_props["price"].mean()) if not df_props.empty else 0.0

    # Transactions du mois en cours
    now = datetime.utcnow()
    start_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    if not df_txns.empty:
        df_txns["transaction_date"] = pd.to_datetime(df_txns["transaction_date"])
        ventes_mois = len(df_txns[df_txns["transaction_date"] >= start_of_month])
        ca_mois = float(df_txns[df_txns["transaction_date"] >= start_of_month]["price_sold"].sum())
    else:
        ventes_mois = 0
        ca_mois = 0.0

    # Taux de conversion
    taux_conversion = round((biens_vendus / total_biens * 100), 1) if total_biens > 0 else 0.0

    return {
        "total_biens": total_biens,
        "biens_disponibles": biens_disponibles,
        "biens_vendus": biens_vendus,
        "biens_en_cours": biens_en_cours,
        "taux_conversion": taux_conversion,
        "ca_total": round(ca_total, 2),
        "ca_mois": round(ca_mois, 2),
        "ventes_mois": ventes_mois,
        "prix_moyen": round(prix_moyen, 2),
    }


# ---------------------------------------------------------------------------
# BY CITY — Stats et prédictions par ville
# ---------------------------------------------------------------------------
@router.get("/by-city")
def get_stats_by_city(
    db: Session = Depends(get_db),
    current_user: core_models.User = Depends(get_current_user),
) -> List[Dict[str, Any]]:
    """
    Analyse Pandas par ville : prix moyen, nombre de biens, CA, tendance.
    Identifie les zones 'chaudes' (forte demande).
    """
    _require_staff(current_user)

    properties = db.query(core_models.Property).all()
    transactions = db.query(core_models.Transaction).all()

    if not properties:
        return []

    df_props = pd.DataFrame([{
        "city": p.city,
        "price": p.price,
        "area": p.area,
        "status": p.status.value,
        "type": p.type.value,
    } for p in properties])

    if transactions:
        df_txns = pd.DataFrame([{
            "property_id": t.property_id,
            "price_sold": t.price_sold,
        } for t in transactions])
        # Joindre avec la ville
        df_props_light = pd.DataFrame([{"id": p.id, "city": p.city} for p in properties])
        df_txns = df_txns.merge(df_props_light, left_on="property_id", right_on="id", how="left")
        ca_by_city = df_txns.groupby("city")["price_sold"].sum().reset_index()
        ca_by_city.columns = ["city", "ca"]
    else:
        ca_by_city = pd.DataFrame(columns=["city", "ca"])

    # Grouper par ville
    stats = df_props.groupby("city").agg(
        total_biens=("city", "count"),
        prix_moyen=("price", "mean"),
        prix_m2_moyen=("area", lambda x: (df_props.loc[x.index, "price"] / x).mean()),
        biens_vendus=("status", lambda x: (x == "sold").sum()),
    ).reset_index()

    # Fusionner avec le CA
    if not ca_by_city.empty:
        stats = stats.merge(ca_by_city, on="city", how="left").fillna(0)
    else:
        stats["ca"] = 0

    # Score de popularité (0-100) basé sur le taux de vente
    stats["taux_vente"] = (stats["biens_vendus"] / stats["total_biens"] * 100).round(1)
    stats["score_popularite"] = (stats["taux_vente"] / stats["taux_vente"].max() * 100).round(0).astype(int) if stats["taux_vente"].max() > 0 else 0

    # Trier par CA décroissant
    stats = stats.sort_values("ca", ascending=False)

    return stats.to_dict(orient="records")


# ---------------------------------------------------------------------------
# MONTHLY TREND — Évolution mensuelle des ventes (12 derniers mois)
# ---------------------------------------------------------------------------
@router.get("/monthly-trend")
def get_monthly_trend(
    db: Session = Depends(get_db),
    current_user: core_models.User = Depends(get_current_user),
) -> List[Dict[str, Any]]:
    """
    Retourne l'évolution mensuelle des ventes sur 12 mois.
    Pandas calcule aussi une ligne de tendance linéaire (prévision).
    """
    _require_staff(current_user)

    transactions = db.query(core_models.Transaction).all()
    if not transactions:
        return []

    df = pd.DataFrame([{
        "price_sold": t.price_sold,
        "date": t.transaction_date,
    } for t in transactions])

    df["date"] = pd.to_datetime(df["date"])
    df["month"] = df["date"].dt.to_period("M")

    # 12 derniers mois
    now = pd.Timestamp.now()
    start = now - pd.DateOffset(months=11)
    df = df[df["date"] >= start]

    monthly = df.groupby("month").agg(
        nb_ventes=("price_sold", "count"),
        ca=("price_sold", "sum"),
    ).reset_index()
    monthly["month"] = monthly["month"].astype(str)
    monthly["ca"] = monthly["ca"].round(2)

    # Tendance linéaire (régression simple) sur le CA
    if len(monthly) >= 3:
        x = np.arange(len(monthly))
        y = monthly["ca"].values
        coeffs = np.polyfit(x, y, 1)
        monthly["tendance"] = (coeffs[0] * x + coeffs[1]).round(2).tolist()
    else:
        monthly["tendance"] = monthly["ca"].tolist()

    return monthly.to_dict(orient="records")


# ---------------------------------------------------------------------------
# TOP PROPERTIES — Biens les plus populaires (prédiction de vente)
# ---------------------------------------------------------------------------
@router.get("/top-properties")
def get_top_properties(
    db: Session = Depends(get_db),
    current_user: core_models.User = Depends(get_current_user),
) -> List[Dict[str, Any]]:
    """
    Identifie les biens avec le meilleur potentiel de vente.
    Score calculé par Pandas selon : prix/m², ville, type de bien.
    """
    _require_staff(current_user)

    props = db.query(core_models.Property).filter(
        core_models.Property.status != core_models.PropertyStatus.SOLD
    ).all()

    if not props:
        return []

    df = pd.DataFrame([{
        "id": p.id,
        "title": p.title,
        "city": p.city,
        "price": p.price,
        "area": p.area,
        "type": p.type.value,
        "status": p.status.value,
        "transaction_mode": p.transaction_mode.value,
    } for p in props])

    df["prix_m2"] = df["price"] / df["area"]

    # Score : biens avec prix_m2 dans la médiane de leur ville → meilleur potentiel
    medians = df.groupby("city")["prix_m2"].median().to_dict()
    df["mediane_ville"] = df["city"].map(medians)
    df["ecart_mediane"] = abs(df["prix_m2"] - df["mediane_ville"]) / df["mediane_ville"]

    # Score de 0 à 100 : plus l'écart est faible, plus le bien est "au bon prix"
    max_ecart = df["ecart_mediane"].max() if df["ecart_mediane"].max() > 0 else 1
    df["score_vente"] = ((1 - df["ecart_mediane"] / max_ecart) * 100).round(0).astype(int)

    top = df.nlargest(10, "score_vente")[["id", "title", "city", "price", "area", "type", "status", "transaction_mode", "prix_m2", "score_vente"]]
    top["prix_m2"] = top["prix_m2"].round(0)

    return top.to_dict(orient="records")


# ---------------------------------------------------------------------------
# RECENT ACTIVITY — 20 derniers événements
# ---------------------------------------------------------------------------
@router.get("/recent-activity")
def get_recent_activity(
    db: Session = Depends(get_db),
    current_user: core_models.User = Depends(get_current_user),
) -> List[Dict[str, Any]]:
    """Retourne les 20 dernières transactions et biens ajoutés, triés par date."""
    _require_staff(current_user)

    # Dernières transactions
    txns = db.query(core_models.Transaction).order_by(
        core_models.Transaction.transaction_date.desc()
    ).limit(10).all()

    # Derniers biens ajoutés
    props = db.query(core_models.Property).order_by(
        core_models.Property.created_at.desc()
    ).limit(10).all()

    activity = []

    for t in txns:
        activity.append({
            "type": "transaction",
            "label": f"Bien vendu — {t.property.title if t.property else 'N/A'}",
            "detail": f"{t.property.city if t.property else ''} · {t.price_sold:,.0f} €",
            "date": t.transaction_date.isoformat() if t.transaction_date else None,
            "color": "green",
        })

    for p in props:
        activity.append({
            "type": "bien_ajoute",
            "label": f"Nouveau bien — {p.title}",
            "detail": f"{p.city} · {p.price:,.0f} €",
            "date": p.created_at.isoformat() if p.created_at else None,
            "color": "blue",
        })

    # Trier par date décroissante
    activity.sort(key=lambda x: x["date"] or "", reverse=True)
    return activity[:20]


# ---------------------------------------------------------------------------
# EXPORT — Téléchargement CSV ou Excel de toutes les données analytiques
# ---------------------------------------------------------------------------
@router.get("/export")
def export_analytics(
    format: str = "excel",          # "csv" ou "excel"
    db: Session = Depends(get_db),
    current_user: core_models.User = Depends(get_current_user),
):
    """
    Génère et télécharge un fichier d'analyse complet :
    - CSV  : fichier unique avec les stats par ville
    - Excel: classeur multi-onglets (KPIs, Villes, Tendance mensuelle, Top biens)
    
    Accès réservé Direction et Commercial.
    """
    from fastapi.responses import StreamingResponse
    import io

    _require_staff(current_user)

    # ── Collecte des données ──────────────────────────────────────────────────
    properties = db.query(core_models.Property).all()
    transactions = db.query(core_models.Transaction).all()

    # DataFrame propriétés
    if properties:
        df_props = pd.DataFrame([{
            "ID": p.id,
            "Titre": p.title,
            "Ville": p.city,
            "Prix (€)": p.price,
            "Surface (m²)": p.area,
            "Prix/m² (€)": round(p.price / p.area, 0) if p.area else 0,
            "Type": p.type.value,
            "Mode": "Vente" if p.transaction_mode.value == "sale" else "Location",
            "Statut": {"available": "Disponible", "under_offer": "Sous offre", "sold": "Vendu"}.get(p.status.value, p.status.value),
            "Date ajout": p.created_at.strftime("%d/%m/%Y") if p.created_at else "",
        } for p in properties])
    else:
        df_props = pd.DataFrame()

    # DataFrame transactions
    if transactions:
        df_txns_raw = pd.DataFrame([{
            "ID transaction": t.id,
            "Bien": t.property.title if t.property else "N/A",
            "Ville": t.property.city if t.property else "",
            "Prix de vente (€)": t.price_sold,
            "Prix initial (€)": t.property.price if t.property else 0,
            "Négociation (%)": round((1 - t.price_sold / t.property.price) * 100, 1) if t.property and t.property.price else 0,
            "Date": t.transaction_date.strftime("%d/%m/%Y") if t.transaction_date else "",
        } for t in transactions])
        df_txns_raw["Date_dt"] = pd.to_datetime([t.transaction_date for t in transactions])
    else:
        df_txns_raw = pd.DataFrame()

    # Stats par ville
    if not df_props.empty:
        df_villes = df_props.groupby("Ville").agg(
            Total_biens=("ID", "count"),
            Prix_moyen=("Prix (€)", "mean"),
            Prix_m2_moyen=("Prix/m² (€)", "mean"),
            Biens_disponibles=("Statut", lambda x: (x == "Disponible").sum()),
            Biens_vendus=("Statut", lambda x: (x == "Vendu").sum()),
        ).reset_index()
        if not df_txns_raw.empty:
            ca_city = df_txns_raw.groupby("Ville")["Prix de vente (€)"].sum().reset_index()
            ca_city.columns = ["Ville", "CA Total (€)"]
            df_villes = df_villes.merge(ca_city, on="Ville", how="left").fillna(0)
        else:
            df_villes["CA Total (€)"] = 0
        df_villes["Taux vente (%)"] = (df_villes["Biens_vendus"] / df_villes["Total_biens"] * 100).round(1)
        df_villes = df_villes.rename(columns={
            "Total_biens": "Total biens",
            "Prix_moyen": "Prix moyen (€)",
            "Prix_m2_moyen": "Prix/m² moyen (€)",
            "Biens_disponibles": "Disponibles",
            "Biens_vendus": "Vendus",
        })
        df_villes = df_villes.sort_values("CA Total (€)", ascending=False)
        df_villes[["Prix moyen (€)", "Prix/m² moyen (€)", "CA Total (€)"]] = df_villes[["Prix moyen (€)", "Prix/m² moyen (€)", "CA Total (€)"]].round(0)
    else:
        df_villes = pd.DataFrame()

    # Tendance mensuelle (12 mois)
    if not df_txns_raw.empty:
        df_txns_raw["Mois"] = df_txns_raw["Date_dt"].dt.to_period("M").astype(str)
        df_tendance = df_txns_raw.groupby("Mois").agg(
            Nb_ventes=("Prix de vente (€)", "count"),
            CA=("Prix de vente (€)", "sum"),
        ).reset_index()
        df_tendance = df_tendance.rename(columns={"Nb_ventes": "Nb ventes", "CA": "CA mensuel (€)"})
        df_tendance["CA mensuel (€)"] = df_tendance["CA mensuel (€)"].round(0)
        df_tendance = df_tendance.sort_values("Mois")
    else:
        df_tendance = pd.DataFrame()

    # KPIs résumé
    now = datetime.utcnow()
    start_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    ventes_mois = 0
    ca_mois = 0.0
    ca_total = 0.0
    if not df_txns_raw.empty:
        ca_total = float(df_txns_raw["Prix de vente (€)"].sum())
        mois_mask = df_txns_raw["Date_dt"] >= pd.Timestamp(start_month)
        ventes_mois = int(mois_mask.sum())
        ca_mois = float(df_txns_raw.loc[mois_mask, "Prix de vente (€)"].sum())

    total_biens = len(properties)
    biens_vendus = sum(1 for p in properties if p.status.value == "sold")
    biens_dispo = sum(1 for p in properties if p.status.value == "available")
    taux_conv = round(biens_vendus / total_biens * 100, 1) if total_biens > 0 else 0

    df_kpis = pd.DataFrame([
        {"Indicateur": "Total biens", "Valeur": total_biens},
        {"Indicateur": "Biens disponibles", "Valeur": biens_dispo},
        {"Indicateur": "Biens vendus", "Valeur": biens_vendus},
        {"Indicateur": "Taux de conversion (%)", "Valeur": taux_conv},
        {"Indicateur": "CA total (€)", "Valeur": round(ca_total, 2)},
        {"Indicateur": "CA ce mois (€)", "Valeur": round(ca_mois, 2)},
        {"Indicateur": "Ventes ce mois", "Valeur": ventes_mois},
        {"Indicateur": "Date d'export", "Valeur": now.strftime("%d/%m/%Y %H:%M")},
    ])

    filename_date = now.strftime("%Y%m%d_%H%M")

    # ── Export CSV ────────────────────────────────────────────────────────────
    if format.lower() == "csv":
        # Pour CSV on exporte les stats par ville (le plus utile en CSV simple)
        buf = io.StringIO()
        buf.write(f"# Export Ymmo Analytics — {now.strftime('%d/%m/%Y %H:%M')}\n")
        buf.write(f"# KPIs : {total_biens} biens | CA total : {ca_total:,.0f} € | Taux conversion : {taux_conv}%\n\n")
        if not df_villes.empty:
            df_villes.to_csv(buf, index=False, sep=";", encoding="utf-8")
        content = buf.getvalue()
        buf.close()
        return StreamingResponse(
            iter([content.encode("utf-8-sig")]),   # BOM pour Excel francophone
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=ymmo_analytics_{filename_date}.csv"},
        )

    # ── Export Excel ──────────────────────────────────────────────────────────
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        # Onglet 1 : KPIs
        df_kpis.to_excel(writer, sheet_name="KPIs", index=False)

        # Onglet 2 : Stats par ville
        if not df_villes.empty:
            df_villes.to_excel(writer, sheet_name="Par ville", index=False)

        # Onglet 3 : Tendance mensuelle
        if not df_tendance.empty:
            df_tendance.to_excel(writer, sheet_name="Tendance mensuelle", index=False)

        # Onglet 4 : Transactions détail
        if not df_txns_raw.empty:
            df_txns_raw.drop(columns=["Date_dt"], errors="ignore").to_excel(
                writer, sheet_name="Transactions", index=False
            )

        # Onglet 5 : Catalogue biens
        if not df_props.empty:
            df_props.to_excel(writer, sheet_name="Catalogue biens", index=False)

        # ── Mise en forme des onglets ──
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        header_fill = PatternFill("solid", fgColor="1E3A5F")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        alt_fill    = PatternFill("solid", fgColor="F0F4F8")
        border_side = Side(style="thin", color="D1D5DB")
        cell_border = Border(bottom=Border(bottom=border_side).bottom)

        for sheet in writer.sheets.values():
            # En-têtes
            for cell in sheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            # Alternance de couleurs de lignes
            for i, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                for cell in row:
                    if i % 2 == 0:
                        cell.fill = alt_fill
                    cell.alignment = Alignment(vertical="center")
            # Ajuster la largeur des colonnes automatiquement
            for col_idx, col in enumerate(sheet.columns, start=1):
                max_len = max((len(str(cell.value or "")) for cell in col), default=10)
                sheet.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)
            # Freeze la première ligne
            sheet.freeze_panes = "A2"

    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=ymmo_analytics_{filename_date}.xlsx"},
    )
